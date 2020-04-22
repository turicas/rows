# coding: utf-8

# Copyright 2014-2019 Álvaro Justen <https://github.com/turicas/rows/>

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Lesser General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Lesser General Public License for more details.

#    You should have received a copy of the GNU Lesser General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

from __future__ import unicode_literals

from decimal import Decimal
from io import BytesIO, UnsupportedOperation
from numbers import Number

from openpyxl import Workbook, load_workbook
from openpyxl.cell.read_only import EmptyCell

from rows import fields
from rows.plugins.utils import create_table, prepare_to_export
from rows.utils import Source


def _cell_to_python(cell):
    """Convert a PyOpenXL's `Cell` object to the corresponding Python object."""
    data_type, value = cell.data_type, cell.value

    if type(cell) is EmptyCell:
        return None
    elif data_type == "f" and value == "=TRUE()":
        return True
    elif data_type == "f" and value == "=FALSE()":
        return False

    elif cell.number_format.lower() == "yyyy-mm-dd":
        return str(value).split(" 00:00:00")[0]
    elif cell.number_format.lower() == "yyyy-mm-dd hh:mm:ss":
        return str(value).split(".")[0]

    elif cell.number_format.endswith("%") and isinstance(value, Number):
        value = Decimal(str(value))
        return "{:%}".format(value)

    elif value is None:
        return ""
    else:
        return value


def sheet_names(filename_or_fobj, workbook_kwargs=None):
    # TODO: setup/teardown must be methods of a class so we can reuse them
    workbook_kwargs = workbook_kwargs or {}
    workbook_kwargs["read_only"] = workbook_kwargs.get("read_only", True)

    workbook = load_workbook(filename_or_fobj, **workbook_kwargs)
    result = workbook.sheetnames
    workbook.close()

    return result


def import_from_xlsx(
    filename_or_fobj,
    sheet_name=None,
    sheet_index=0,
    start_row=None,
    start_column=None,
    end_row=None,
    end_column=None,
    workbook_kwargs=None,
    *args,
    **kwargs
):
    """Return a rows.Table created from imported XLSX file.

    workbook_kwargs will be passed to openpyxl.load_workbook
    """

    workbook_kwargs = workbook_kwargs or {}
    workbook_kwargs["read_only"] = workbook_kwargs.get("read_only", True)

    workbook = load_workbook(filename_or_fobj, **workbook_kwargs)
    if sheet_name is None:
        sheet_name = workbook.sheetnames[sheet_index]
    sheet = workbook[sheet_name]

    # The openpyxl library reads rows and columns starting from 1 and ending on
    # sheet.max_row/max_col. rows uses 0-based indexes (from 0 to N - 1), so we
    # need to adjust the ranges accordingly.
    min_row = sheet.min_row - 1 if sheet.min_row is not None else None
    min_column = sheet.min_column - 1 if sheet.min_column is not None else None
    max_row = sheet.max_row - 1 if sheet.max_row is not None else None
    max_column = sheet.max_column - 1 if sheet.max_column is not None else None
    # TODO: consider adding a parameter `ignore_padding=True` and when it's
    # True, consider `start_row` starting from `sheet.min_row` and
    # `start_column` starting from `sheet.min_col`.
    start_row = start_row if start_row is not None else min_row
    end_row = end_row if end_row is not None else max_row
    start_column = start_column if start_column is not None else min_column
    end_column = end_column if end_column is not None else max_column
    table_rows = []
    is_empty = lambda row: all(cell is None for cell in row)
    selected_rows = sheet.iter_rows(
        min_row=start_row + 1 if start_row is not None else None,
        max_row=end_row + 1 if end_row is not None else None,
        min_col=start_column + 1 if start_column is not None else None,
        max_col=end_column + 1 if end_column is not None else None,
    )
    for row in selected_rows:
        row = [_cell_to_python(cell) for cell in row]
        if not is_empty(row):
            table_rows.append(row)

    source = Source.from_file(filename_or_fobj, plugin_name="xlsx")
    source.fobj.close()
    # TODO: pass a parameter to Source.from_file so it won't open the file
    metadata = {"imported_from": "xlsx", "source": source, "name": sheet_name}
    return create_table(table_rows, meta=metadata, *args, **kwargs)


FORMATTING_STYLES = {
    fields.DateField: "YYYY-MM-DD",
    fields.DatetimeField: "YYYY-MM-DD HH:MM:SS",
    fields.PercentField: "0.00%",
}


def _python_to_cell(field_types):
    def convert_value(field_type, value):

        number_format = FORMATTING_STYLES.get(field_type, None)

        if field_type not in (
            fields.BoolField,
            fields.DateField,
            fields.DatetimeField,
            fields.DecimalField,
            fields.FloatField,
            fields.IntegerField,
            fields.PercentField,
            fields.TextField,
        ):
            # BinaryField, DatetimeField, JSONField or unknown
            value = field_type.serialize(value)

        return value, number_format

    def convert_row(row):
        return [
            convert_value(field_type, value)
            for field_type, value in zip(field_types, row)
        ]

    return convert_row


def define_sheet_name(existing_names):
    for counter in range(1, 1024 * 1024):
        new_name = f"Sheet{counter}"
        if new_name not in existing_names:
            return new_name


def is_existing_spreadsheet(source):
    if source.uri is not None:  # filename was given
        if not source.uri.exists():
            # TODO: if file doesn't exist and we open with mode="a+b" it will
            # be created and therefore this `if` won't be `True`.
            return False

    fobj = source.fobj
    fobj.seek(0)
    try:
        data = fobj.read(1024)
    except UnsupportedOperation:
        # File in write-only mode: so it's a new file
        return False
    else:
        fobj.seek(0)
        return data[:2] == b"PK"  # XXX: any zip file will return `True`


def export_to_xlsx(table, filename_or_fobj=None, sheet_name=None, *args, **kwargs):
    """Export the rows.Table to XLSX file and return the saved file."""

    return_result = False
    if filename_or_fobj is None:
        filename_or_fobj = BytesIO()
        return_result = True
    source = Source.from_file(filename_or_fobj, mode="a+b", plugin_name="xlsx")

    if is_existing_spreadsheet(source):
        workbook = load_workbook(filename_or_fobj)
        if sheet_name is None:
            sheet_name = define_sheet_name(workbook.sheetnames)
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet_name = sheet_name or "Sheet1"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    prepared_table = prepare_to_export(table, *args, **kwargs)

    # Write header
    field_names = next(prepared_table)
    for col_index, field_name in enumerate(field_names):
        cell = sheet.cell(row=1, column=col_index + 1)
        cell.value = field_name

    # Write sheet rows
    _convert_row = _python_to_cell(list(map(table.fields.get, field_names)))
    for row_index, row in enumerate(prepared_table, start=1):
        for col_index, (value, number_format) in enumerate(_convert_row(row)):
            cell = sheet.cell(row=row_index + 1, column=col_index + 1)
            cell.value = value
            if number_format is not None:
                cell.number_format = number_format

    source.fobj.seek(0)
    if source.uri is not None:
        # For some reason the `ZipFile` inside
        # `openpyxl.workbook.workbook.save_workbook` was not creating the
        # contents correctly when a fobj is passed, so filename is forced.
        workbook.save(source.uri)
    else:
        workbook.save(source.fobj)
    source.fobj.flush()

    if return_result:
        source.fobj.seek(0)
        result = source.fobj.read()
    else:
        result = source.fobj

    if source.should_close:
        source.fobj.close()

    return result
